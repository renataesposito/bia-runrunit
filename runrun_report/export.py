import io
import os
import copy
import unicodedata
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "yesh_nuclea_template.xlsx")

_BLUE_FILL  = PatternFill("solid", fgColor="1F3864")
_GREEN_FILL = PatternFill("solid", fgColor="198754")
_BOLD_WHITE = Font(bold=True, color="FFFFFF")

def _normalize(text):
    """Normaliza texto para comparação: minúsculas, sem acentos, sem espaços extras."""
    if not text or not isinstance(text, str):
        return ""
    text = text.lower().strip()
    # Remove acentos
    text = "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if unicodedata.category(c) != "Mn"
    )
    # Remove caracteres não alfanuméricos simples (mantém letras e números)
    text = re.sub(r"[^a-z0-9]", "", text)
    return text

def _style_header(ws, fill=None):
    fill = fill or _BLUE_FILL
    for cell in ws[1]:
        cell.fill = fill
        cell.font = _BOLD_WHITE
        cell.alignment = Alignment(horizontal="center")

def _auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        values = [str(cell.value) if cell.value is not None else '' for cell in col]
        width  = min(max_w, max(min_w, max((len(v) for v in values), default=min_w) + 2))
        ws.column_dimensions[col[0].column_letter].width = width

def _copy_cell_style(source_cell, target_cell):
    """Copia a formatação de uma célula para outra."""
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)

def _find_kpi_cell(ws, label):
    """Encontra a célula que contém um texto específico na aba de KPIs."""
    norm_label = _normalize(label)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if norm_label in _normalize(cell.value):
                    return cell
    return None

def _find_table_header_row(ws, expected_headers):
    """
    Procura a linha que contém os cabeçalhos esperados.
    Retorna (row_idx, col_mapping) onde col_mapping é um dict: {nome_coluna_normalizado: col_idx}
    """
    expected_norm = {h: _normalize(h) for h in expected_headers}
    
    # Aumentando range de busca para 100 linhas caso o template tenha muito conteúdo inicial
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100), start=1):
        found_headers = {}
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_norm = _normalize(cell.value)
                if not val_norm: continue
                
                for original_name, norm_exp in expected_norm.items():
                    # Match flexível: um contido no outro
                    if norm_exp in val_norm or val_norm in norm_exp:
                        found_headers[original_name] = cell.column
        
        # Se encontrou pelo menos 2 cabeçalhos (para tabelas pequenas) ou 40% dos esperados
        if len(found_headers) >= max(2, len(expected_headers) * 0.4):
            return row_idx, found_headers
            
    return None, {}

def _clear_and_fill_table(ws, df, start_row, col_mapping):
    """
    Limpa as linhas de dados da tabela (preservando o template) e insere os novos dados.
    Utiliza uma abordagem que evita deletar linhas para não quebrar âncoras de imagens.
    """
    # 1. Armazena o estilo da primeira linha de dados do template (logo abaixo do cabeçalho)
    reference_styles = {}
    for original_name, col_idx in col_mapping.items():
        ref_cell = ws.cell(row=start_row + 1, column=col_idx)
        # Salva uma cópia do estilo para aplicar nas novas linhas
        reference_styles[col_idx] = {
            'font': copy.copy(ref_cell.font),
            'border': copy.copy(ref_cell.border),
            'fill': copy.copy(ref_cell.fill),
            'number_format': copy.copy(ref_cell.number_format),
            'protection': copy.copy(ref_cell.protection),
            'alignment': copy.copy(ref_cell.alignment)
        }
        
    # 2. Limpa TODOS os dados existentes a partir do start_row + 1 até o fim da planilha
    # Limpamos uma área grande para garantir que dados de exemplo sumam
    max_row = max(ws.max_row, start_row + 500) # Limpa pelo menos 500 linhas
    max_col = max(ws.max_column, 20)            # Limpa pelo menos 20 colunas
    
    for r in range(start_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None
        
    # 3. Insere os novos dados (se houver)
    if df is not None and not df.empty:
        # Normaliza nomes de colunas do DF para match seguro
        df_cols_norm = {c: _normalize(c) for c in df.columns}
        
        for i, (_, row_data) in enumerate(df.iterrows()):
            current_row_idx = start_row + 1 + i
            
            for original_name, col_idx in col_mapping.items():
                cell = ws.cell(row=current_row_idx, column=col_idx)
                norm_target = _normalize(original_name)
                
                val = None
                for df_col_orig, df_col_norm in df_cols_norm.items():
                    if df_col_norm == norm_target or norm_target in df_col_norm or df_col_norm in norm_target:
                        val = row_data[df_col_orig]
                        break
                        
                cell.value = val
                
                # Aplica estilo de referência
                if col_idx in reference_styles:
                    style = reference_styles[col_idx]
                    cell.font = copy.copy(style['font'])
                    cell.border = copy.copy(style['border'])
                    cell.fill = copy.copy(style['fill'])
                    cell.number_format = copy.copy(style['number_format'])
                    cell.protection = copy.copy(style['protection'])
                    cell.alignment = copy.copy(style['alignment'])

def gerar_excel(escopo_real: pd.DataFrame, entregas: pd.DataFrame, kpis: dict) -> bytes:
    # Verifica se o arquivo de template existe. Se não existir, usa o fallback que gera do zero.
    if not os.path.exists(TEMPLATE_PATH):
        return _gerar_excel_fallback(escopo_real, entregas, kpis)
        
    wb = load_workbook(TEMPLATE_PATH)

    # --- Aba 1: Resumo KPIs ---
    ws_kpi = wb["Resumo KPIs"] if "Resumo KPIs" in wb.sheetnames else wb.active
    if ws_kpi:
        # Cálculo da porcentagem anual conforme solicitado (Realizado / Total do Contrato)
        total_realizado = kpis.get("total_realizado", 0)
        total_contrato = kpis.get("total_contrato", 0)
        pct_anual = round(100 * total_realizado / total_contrato, 1) if total_contrato else 0
        
        kpi_mapping = {
            "Meses decorridos": f"{int(kpis.get('meses_decorridos', 0))} meses",
            "Entregas Previstas": kpis.get("total_previsto", 0),
            "Entregas Realizadas": total_realizado,
            "% de Realização": f"{pct_anual}%"
        }
        
        for label, value in kpi_mapping.items():
            cell = _find_kpi_cell(ws_kpi, label)
            if cell:
                target_cell = ws_kpi.cell(row=cell.row, column=cell.column + 1)
                target_cell.value = value

    # --- Aba 2: Escopo x Realizado ---
    if "Escopo x Realizado" in wb.sheetnames:
        ws_esc = wb["Escopo x Realizado"]
        # Mapeamento de colunas conforme solicitado: "Escopo Ano" deve ser o "qtd_ano"
        cols_esc = ["grupo", "entregavel", "qtd_mes", "qtd_ano", "previsto_acumulado", "realizado"]
        df_esc = escopo_real[[c for c in cols_esc if c in escopo_real.columns]].copy()
        
        # Nomes das colunas para busca no template (incluindo "Escopo Ano" conforme solicitado)
        display_cols = ["Grupo", "Entregável", "Qtd/Mês", "Escopo Ano", "Previsto Acumulado", "Realizado"]
        df_esc.columns = display_cols[:len(df_esc.columns)]
        
        header_row_idx, col_mapping = _find_table_header_row(ws_esc, display_cols)
        
        if header_row_idx:
            _clear_and_fill_table(ws_esc, df_esc, header_row_idx, col_mapping)
            _auto_width(ws_esc)

    # --- Aba 3: Histórico Entregas ---
    if "Histórico Entregas" in wb.sheetnames:
        ws_hist = wb["Histórico Entregas"]
        
        # Prepara os dados (mesmo que entregas esteja vazio, precisamos limpar a aba)
        df_hist = pd.DataFrame()
        if not entregas.empty:
            slug_to_name = escopo_real.set_index("slug")["entregavel"].to_dict()
            df_hist = entregas.copy()
            
            if "entregavel" not in df_hist.columns:
                df_hist.insert(4, "entregavel", df_hist["scope_slug"].map(slug_to_name).fillna(df_hist["hashtag"]))
                
            # Colunas solicitadas: 'Mês/Ano', 'Grupo', 'Projeto', 'Entregável' e 'Quantidade'
            cols_hist = ["mes_ano", "grupo", "projeto", "entregavel", "quantidade"]
            df_hist = df_hist[[c for c in cols_hist if c in df_hist.columns]]
            
            display_cols_hist = ["Mês/Ano", "Grupo", "Projeto", "Entregável", "Quantidade"]
            df_hist.columns = display_cols_hist[:len(df_hist.columns)]
        else:
            display_cols_hist = ["Mês/Ano", "Grupo", "Projeto", "Entregável", "Quantidade"]
        
        header_row_idx, col_mapping = _find_table_header_row(ws_hist, display_cols_hist)
        
        if header_row_idx:
            # Chama sempre, para limpar o template mesmo que não haja novos dados
            _clear_and_fill_table(ws_hist, df_hist, header_row_idx, col_mapping)
            _auto_width(ws_hist)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _gerar_excel_fallback(escopo_real: pd.DataFrame, entregas: pd.DataFrame, kpis: dict) -> bytes:
    """Geração de excel do zero caso o template não exista."""
    wb = Workbook()

    # Aba 1: KPIs
    ws_kpi = wb.active
    ws_kpi.title = "Resumo KPIs"
    
    total_realizado = kpis.get("total_realizado", 0)
    total_contrato = kpis.get("total_contrato", 0)
    pct_anual = round(100 * total_realizado / total_contrato, 1) if total_contrato else 0

    for row in [
        ["Indicador", "Valor"],
        ["Meses decorridos", f"{int(kpis.get('meses_decorridos', 0))} meses"],
        ["Entregas Previstas",     kpis.get("total_previsto", 0)],
        ["Entregas Realizadas",                total_realizado],
        ["% de Realização",                    f"{pct_anual}%"],
    ]:
        ws_kpi.append(row)
    _style_header(ws_kpi)
    _auto_width(ws_kpi)

    # Aba 2: Escopo x Realizado
    ws_esc = wb.create_sheet("Escopo x Realizado")
    cols_esc = ["grupo", "entregavel", "qtd_mes", "qtd_ano", "previsto_acumulado", "realizado"]
    df_esc = escopo_real[[c for c in cols_esc if c in escopo_real.columns]].copy()
    df_esc.columns = ["Grupo", "Entregável", "Qtd/Mês", "Escopo Ano",
                      "Previsto Acumulado", "Realizado"][:len(df_esc.columns)]
    for row in dataframe_to_rows(df_esc, index=False, header=True):
        ws_esc.append(row)
    _style_header(ws_esc)
    _auto_width(ws_esc)

    # Aba 3: Histórico de Entregas
    ws_hist = wb.create_sheet("Histórico Entregas")
    if not entregas.empty:
        slug_to_name = escopo_real.set_index("slug")["entregavel"].to_dict()
        df_hist = entregas.copy()
        if "entregavel" not in df_hist.columns:
            df_hist.insert(4, "entregavel", df_hist["scope_slug"].map(slug_to_name).fillna(df_hist["hashtag"]))
        
        cols_hist = ["mes_ano", "grupo", "projeto", "entregavel", "quantidade"]
        df_hist = df_hist[[c for c in cols_hist if c in df_hist.columns]]
        df_hist.columns = ["Mês/Ano", "Grupo", "Projeto",
                           "Entregável", "Quantidade"][:len(df_hist.columns)]
        for row in dataframe_to_rows(df_hist, index=False, header=True):
            ws_hist.append(row)
        _style_header(ws_hist)
        _auto_width(ws_hist)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
